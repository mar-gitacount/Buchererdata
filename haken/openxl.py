import openpyxl
import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font

current_directory = os.getcwd()
# エクセルファイルのパスを指定
# excel_file_path = r'C:\Users\01794\Desktop\仕事でつかうやつ\Buchererデータ抽出\エクセル\BUCHRER CPOリスト2024.01.22最新.xlsx'
# !エクセルパスは指定しておく
excel_file_path = r"エクセル\BUCHRER CPOリスト2024.03.04訂正.xlsx"
# エクセルファイルの絶対パスを生成
xlsx_file_path = os.path.join(current_directory, excel_file_path)


# 当月シートを取得するための変数を取得

# 現在の日付を取得
current_date = datetime.now()

# 今月の月を取得 (1月なら1、2月なら2、...)
current_month = current_date.month

# "月"という文字列に追加
month_sheet_name = str(current_month) + "月"

# エクセルファイルを開く
workbook = openpyxl.load_workbook(excel_file_path)

# 指定されたシートを取得
sheet = workbook[month_sheet_name]

# A列の値を取得
column_a_values = [cell.value for cell in sheet["A"]]
column_b_values = [cell.value for cell in sheet["B"]]


# ２列目の値が挿入されている最後の行を取得
def get_last_row_with_value(sheet, column=2):
    last_row = None
    for row in range(1, sheet.max_row + 1):
        if (
            sheet.cell(row=row, column=column - 1).value is None
            and sheet.cell(row=row, column=column).value is None
        ):
            # if sheet.cell(row=row, column=column-1).value and sheet.cell(row=row, column=column).value:
            last_row = row
    return last_row


def b_values_get(sheet, array):
    # A列の値を取得
    column_a_values = [cell.value for cell in sheet["A"]]
    column_b_values = [cell.value for cell in sheet["B"]]
    values = []
    values.append = a_value
    flg = True

    while flg:
        for index, a_value in enumerate(column_a_values, start=1):
            # 取得したA列の値が存在しているかつ、となりのセルに値が存在しない場合。
            b_index = column_b_values.index(a_value) + 1
            if (
                a_value in column_b_values
                and sheet.cell(row=index, column=2).value is None
            ):
                sheet.cell(row=index, column=2).value = sheet.cell(
                    row=b_index, column=2
                ).value
            # 値が存在していてかつ隣のセルに値がある場合、A列とB列の値は違う
            elif a_value in column_b_values:
                # B列を配列に退避する
                values.append(sheet.cell(row=b_index, column=2).value)
                # 値を代入する
                sheet.cell(row=index, column=2).value = sheet.cell(
                    row=b_index, column=2
                ).value
                # A列のすぐ隣を開ける
                sheet.cell(row=b_index, column=2).value = ""
            # 値が存在しない場合、新しい行を追加して赤文字にする。
            else:
                sheet.insert_rows(idx=index, amount=1)


# A列の値がB列に存在するかどうかを判定
# A列の値がB列に存在する場合、その位置を表示
value_array = []
for index, a_value in enumerate(column_a_values, start=1):
    column_b_values = [cell.value for cell in sheet["B"]]
    if not a_value:
        continue
    print("---")
    if a_value in column_b_values:
        # if len(int(index)) < 3:
        #     continue
        b_index = column_b_values.index(a_value) + 1

        print(f"A列の値 {a_value} はB列に存在します。位置: A列{index}, B列{b_index}")
        # 隣の値が同じ場合、処理を次の処理へ
        if index == b_index:
            continue
        # A列の隣の値を取得する
        index_b_value = sheet.cell(row=index, column=2).value
        print(f"{index_b_value}はB列から取得した値です!!")
        # A列の隣が空でない場合配列に退避する
        if index_b_value:
            value_array.append(index_b_value)

        # index_b_column = sheet.cell(row=index, column=2).value
        # 取得したB列の値をA列の隣にコピー
        sheet.cell(row=index, column=2).value = sheet.cell(row=b_index, column=2).value

        # 値を取得したもとのセルを空にする。
        sheet.cell(row=b_index, column=2).value = ""
        # sheet.cell(row=b_index, column=2).value = ''
        # 同じ値が存在かつ列が違う場合、キャストしてA列の隣に挿入する。

    else:
        # A列に値が存在しない場合、配列もループしてみて。あった場合、そのとなりの列にいれる
        # となりの列に入れた後、その値をvalue_arrayから削除する。
        # b_index = column_b_values.index(a_value) + 1
        index_b_value = sheet.cell(row=index, column=2).value
        print(f"{index_b_value}はB列から取得した値です!!")

        if index_b_value:
            value_array.append(index_b_value)

        a_value_check_filtered_list = [item for item in value_array if item == a_value]
        if len(a_value_check_filtered_list) == 0:
            # a_value が value_array に存在しない場合の処理
            print(f"{a_value} は value_array に存在しません。")
            # ここに a_value が存在しない場合に行いたい処理を追加
            # ここに a_value が存在する場合に行いたい処理を追加
            print(f"A列の値{a_value}は存在しない、隣にグレーのセルを挿入します。")
            gray_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            sheet.cell(row=index, column=2).value = ""
            # 追加する行
            #  insert_index = index
            # print(f"{index}は追加する行")
            # 指定位置に新しい行を挿入
            # sheet.insert_rows(idx=insert_index, amount=1)
            # sheet.insert_rows(insert_index,amount=1)
            sheet.cell(row=index, column=2).fill = gray_fill
        # sheet.append([None, None])  # 新しい行を追加

        # 下に行をキャスト
        # sheet.insert_rows(idx=index, amount=1)
        else:
            # a_value が value_array に存在する場合の処理
            print(f"{a_value} は value_array に存在します。")
            sheet.cell(row=index, column=2).value = a_value_check_filtered_list[0]
            index_of_element = value_array.index(a_value_check_filtered_list[0])
            # value_array から要素を削除
            value_array.pop(index_of_element)
            continue


# !処理が抜けても配列内に値がある場合、それは存在しない、赤い文字の値。
# for index, a_value in enumerate(column_a_values, start=1):
print(f"{len(value_array)}は配列の長さです!!")
for new_value in value_array:
    lastlow = get_last_row_with_value(sheet, column=2)
    print(f"{new_value}が赤文字")
    if new_value:
        sheet.cell(row=lastlow, column=2).value = new_value
        font_red = Font(color="FF0000")
        sheet.cell(row=lastlow, column=2).font = font_red


column_a_values = [cell.value for cell in sheet["A"]]
column_b_values = [cell.value for cell in sheet["B"]]
# まったく新しい列の値を取得する。
for index, b_value in enumerate(column_b_values, start=1):
    if b_value in column_a_values:
        continue
    elif sheet.cell(row=index, column=1).value is None:
        font_red = Font(color="FF0000")
        sheet.cell(row=index, column=2).font = font_red

    if a_value not in column_b_values:
        print(f"A列の値 {a_value} はB列に存在しません。新しい行を追加します。")

        # 新しい行を追加
        # sheet.insert_rows(idx=index, amount=1)

        # B列の値を新しい行のB列に張り付け
        # sheet.cell(row=index, column=2).value = a_value

        # !赤文字に変更
        # font_red = Font(color="FF0000")
        # sheet.cell(row=index, column=2).font = font_red


workbook.save(excel_file_path)

workbook.close()

# A列とB列の値を比較
# for a_value, b_value in zip(sheet['A'], sheet['B']):
#     if a_value.value == b_value.value:
#         print(f"A列とB列のセルが一致: {a_value.value}")
#     else:
#         print(f"A列とB列のセルが一致しない: A列={a_value.value}, B列={b_value.value}")
#         # セルが一致しない場合、B列の値を走査する


# 結果を表示
# print(column_a_values)


# 例として、A列のセルの値を表示
# for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
#     print(row[0])

# ワークブックを閉じる
workbook.close()


# 結果を表示
# print(result_string)
# 月シートを取得する


# # ワークブックを開く
# workbook = openpyxl.load_workbook(excel_file_path)

# # シートを選択（例: 最初のシートを選択）
# sheet = workbook.active

# # 367行目、A列に値を入力
# sheet.cell(row=367, column=1, value='Your Value Here')

# # ワークブックを保存
# workbook.save(excel_file_path)


# # ここで必要な処理を行う
# # 例: シートのデータを表示
# for row in sheet.iter_rows(values_only=True):
#     print(row)

# # ワークブックを閉じる
# workbook.close()
