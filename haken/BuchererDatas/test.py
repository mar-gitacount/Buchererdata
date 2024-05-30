
from whochedata_sqlite_data_insert import WhocheSqliteDataInsert
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
def main():
    print("test")
    # インスタンスを作成する。
    db_file = "bucherer.db"
    # dbのファイルをわたす。
    whocequeryinstance = WhocheSqliteDataInsert(db_file)
    # テーブル一覧プロパティ
    tablename = whocequeryinstance.table_names
    items = whocequeryinstance.tablesdateill
    print(tablename,"テーブル名")
    # 現在の日付を取得
    today_date = datetime.now().strftime("%Y%m%d")
    file_name = f"ブヘラ{today_date}.xlsx"
  
    # print(items,"アイテム名")
    # テストで値を入れてみる
    # print(items['watch_item'].insert_data("watch_item",1, 2022, "Model 1", "REF123", "Bracelet 1", "Dial 1", "example.com"),"インスタンスたち")
    # items内にすべてのフィールドが格納されている。
    # 一致した場合、その日付のアイテムは除きたい。
    # 
    # jsondataloadmakeexcel.pyに移植する。
    week_items = []
    heder_rows =[]
    # 本日の日付を取得する
    today_date = datetime.now().date()
    # 月
    current_month = today_date.month
    # 変換後
    today_date_stringfomat = today_date.strftime('%Y/%m/%d')
    # 一回目の探索 items['weekly_reports'] weekdateフィールドを探索する　インスタンス　を使う。
    # trueなら配列にいれる？？
    # 本日の値が入っていた場合配列に日付を入れる
    week_date_count = items['weekly_reports'].datacountcheck(today_date_stringfomat,"weekdate")
    week_items.append(today_date_stringfomat) if week_date_count else print("アイテムなし")
    print(f"{week_date_count}は週があるかどうか")
    diff_date = today_date

    while True:
        # 日付で検索、falseの場合、-1日
        # 検索でヒットした場合、その値の月を取得、そのまえの日と比べて月を跨いだら処理終了
        diff_date = diff_date-timedelta(1)
        # 月を取得する  
        diff_date_month = diff_date.month
        # 文字列にする
        diff_date_string = diff_date.strftime('%Y/%m/%d')
        # 月が替わりかつそのアイテムがあればtrueになる。
        week_date_count = items['weekly_reports'].datacountcheck(diff_date_string ,"weekdate")
        
        #アイテムを加える。
        week_items.append(diff_date_string) if week_date_count else print("アイテムなし")
        
        # n日のデータが入り、かつ別月の場合処理を抜ける
        if week_date_count and current_month != diff_date_month:
            break
        if len(week_items) >= 4:
            break
    # 各週をチェックする。
    monthdata = whocequeryinstance.days_diffcheck('weekly_reports',"test",week_items)
    
   
    reversed_week_items = week_items.reverse()
    print(week_items)
    new_index=[""]*len(week_items)
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(week_items)
    else:
        # ファイルが存在する場合は既存のファイルを読み込み
        wb = load_workbook(file_name)
        ws = wb.active

    # 月ごとファイル作成OK
    for id,dates in monthdata.items():
         print(f"呼び出し元→ID: {id}, Dates: {dates}")
         for date in dates:
            index = week_items.index(date)
            print(f"{date}日付 番号:{id}")
            print("-------------")
            new_index.insert(index,id)
        # データを加える
         print(new_index)
         ws.append(new_index)
         new_index=[""]*len(week_items)
    # 一旦エクセル保存する。
    wb.save(file_name)
    # 今週、先週差分チェック
    # 最初と最後のアイテムを抽出する。
    if len(week_items)>=2:
        last_element = week_items[-1]
        second_last_element = week_items[-2]
        week_items = week_items[-2:]
    today_date_stringfomat = today_date.strftime("%Y-%m-%d")


    new_sheet = wb.create_sheet(title=today_date_stringfomat)
    monthdata = whocequeryinstance.days_diffcheck('weekly_reports',"test",week_items)
    
    
    for id,dates in monthdata.items():
        print(f"呼び出し元→ID: {id}, Dates: {dates}")
        for date in dates:
            index = week_items.index(date)
            print(f"{date}日付 番号:{id}")
            print("-------------")
            new_index.insert(index,id)
        new_sheet.append(new_index)
        new_index=[""]*len(week_items)

 
    wb.save(file_name)
    # for item in week_items:
    #     print(f"{week_items}は週ぜんぶ　")
    #     whocequeryinstance.days_diffcheck('weekly_reports',item,week_items)
    #     print(f"{item}は週ごと")
    #     # ここから各週ごとのアイテムをチェックする
    #     # アイテム(n)→週分検索する。

    # 
    for i in range(4):
        print(i)
        print(today_date-timedelta(1))
        today_date = today_date-timedelta(1)
        today_date_stringfomat = today_date.strftime('%Y/%m/%d')
        print(today_date_stringfomat)

    
    dbinsert_datagetnow = datetime.strptime(today_date,'%Y/%m/%d').date()
    print(dbinsert_datagetnow)
    # for i in range(4):
    #     print(i+1)
    #     adddate = i+1
    #     # 七日マイナスする。
    #     # 計算式= n日 - (7*(4-k))
    #     print(adddate)
    #     print(type(dbinsert_datagetnow))
                            
    #     print(type(dbinsert_datagetnow))
    #     print(dbinsert_datagetnow)
    #     dbinsert_datagetnow_minusdate = dbinsert_datagetnow - timedelta(days=7*adddate)
    #     # 月が違う場合処理してbreakする。
    #     if dbinsert_datagetnow.month == dbinsert_datagetnow_minusdate.month:
    #          print("同じ月")
    #     else:
    #          print("違う月、ここでおわり")
    #          print(f"{dbinsert_datagetnow.month}と{dbinsert_datagetnow_minusdate.month}")
    #          # dbinsert_datagetnow_minusdate = dbinsert_datagetnow - timedelta(days=1)
    #          print(dbinsert_datagetnow_minusdate)
                            

    fieldcount = items['watch_item'].fieldcountAllcountcheck()
   
    
    
    
    fieldcount = items['watch_item'].fieldcountAllcountcheck()
    value = [f"test" ,2022, "Model 2", "REF123", "Bracelet 1", "Dial 1", "example.com"]
    jikou = items['watch_item'].insert_data(value)
    weeklydatas = ["2024/05/23","0","4000","1348-535-3"]
    weekinserttest = items['weekly_reports'].insert_data(weeklydatas)


    print(items,"テーブル一覧")
    # テーブル数を確認して自動増分する。
    
 
    # jikou = whocequeryinstance.insertsuper(value)

    # jikou.close_connection()


if __name__ == "__main__":
    main()